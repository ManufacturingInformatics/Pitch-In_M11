from scipy import fftpack
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from datetime import datetime
import pywt
import scaleogram as scg

scg.set_default_wavelet('morl')

# read in the workbook
wb = load_workbook(filename="T14 H3 316 0939.xlsx",data_only=True)
print(f"workbook has sheets called {wb.sheetnames}")
# get data worksheet
print("accessing data worksheet")
data_ws = wb["Data"]
# data is broken up into collections of 5 columns with one columns separating each group
# we're interested in the middle three
# put data into an array
#all_data= np.array([d for d in data_ws.iter_rows(min_row=8,max_col=60,values_only=True)])
# dictionary to hold data
data = {}
# starting column index
mc = 2
# plotting axes
f,ax = plt.subplots()
# iterate over turnable speed groupings
for ttg in range(1,11):
    rpm = 2*ttg
    print(f"attempting to access {rpm} RPM data")
    if rpm==10:
        mc += 1
    print(f"col {mc} {mc+3}")
    # get data for the specific grouping
    dgroup = np.array([d for d in data_ws.iter_rows(min_row=8,min_col=mc,max_col=mc+3,values_only=True)])
    # update column index
    mc += 6
    # add to dictionary
    data[rpm]=dgroup.astype("float64")
    ## plot data
    # find where nans don't occur
    nn = ~ np.isnan(data[rpm][:,2])
    ax.clear()
    scg.cws(data[rpm][:,1][nn],data[rpm][:,2][nn]-data[rpm][:,2][nn].mean(),scales=np.arange(1,150),yscale='log',ax=ax,cmap='jet',cbar=None,ylabel="Period [mins]",xlabel="Time [mins]",title=f"Scaleogram for {rpm} RPM with log scale")
    f.savefig(f"scaleogram_flowrate_TT_{rpm}_logscale.png")

    ax.clear()
    scg.cws(data[rpm][:,1][nn],data[rpm][:,2][nn]-data[rpm][:,2][nn].mean(),scales=np.arange(1,150),ax=ax,cmap='jet',cbar=None,ylabel="Period [mins]",xlabel="Time [mins]",title=f"Scaleogram for {rpm} RPM with log scale")
    f.savefig(f"scaleogram_flowrate_TT_{rpm}.png")

