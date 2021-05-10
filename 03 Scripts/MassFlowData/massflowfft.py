from scipy import fftpack
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt

# read in the workbook
wb = load_workbook(filename="T14 H3 316 0939.xlsx",data_only=True)
print(f"workbook has sheets called {wb.sheetnames}")
# get data worksheet
print("accessing data worksheet")
data_ws = wb["Data"]
# data is broken up into collections of 5 columns with one columns separating each group
# we're interested in the middle three
# put data into an array
#all_data= np.array([d for d in data_ws.iter_rows(min_row=8,max_col=60,values_only=True)])
# dictionary to hold data
data = {}
# starting column index
mc = 2
# plotting axes
f,ax = plt.subplots()
# iterate over turnable speed groupings
for ttg in range(1,11):
    rpm = 2*ttg
    print(f"attempting to access {rpm} RPM data")
    if rpm==10:
        mc += 1
    print(f"col {mc} {mc+3}")
    # get data for the specific grouping
    dgroup = np.array([d for d in data_ws.iter_rows(min_row=8,min_col=mc,max_col=mc+3,values_only=True)])
    # save subdataset
    np.savetxt(f"{rpm}_rpm_data.csv",dgroup.astype("float64"),delimiter=',')
    # update column index
    mc += 6
    # add to dictionary
    data[rpm]=dgroup.astype("float64")
    # plot the data
    ax.clear()
    ax.plot(data[rpm][:,1],data[rpm][:,2],'b-',data[rpm][:,1],data[rpm][:,3],'r-')
    ax.set(xlabel="Time (min)",ylabel="Flow Rate (g/min)",title=f"Flow Rate for TT {rpm} % of 10RPM")
    leg = f.legend(["Raw","Smoothed"])
    f.savefig(f"flowrate_TT_{rpm}.png")
    ## fft analysis
    print("performing fft analysis")
    # remove legend
    leg.remove()
    # compute fourier transform
    mfr_fft = fftpack.fft(data[rpm][:,2])
    # power spectral density
    mfr_psd = np.abs(data[rpm][:,2])**2
    # get frequencies correponding to PSD
    # setting frequency to 1 to 1 min
    fftfreq = fftpack.fftfreq(len(mfr_psd),1./60.)
    # only getting the positive parts as they correspond to the real signal
    i = fftfreq>0
    ## plot
    ax.clear()
    ax.plot(fftfreq[i],10*np.log10(mfr_psd[i]))
    ax.set(xlabel="Freq (1/min)",ylabel="Power Spectral Density (dB)",title=f"Power Spectral Density TT {rpm} % of 10RPM")
    f.savefig(f"fft_flowrate_TT_{rpm}.png")

    ax.clear()
    ax.plot(fftfreq[i],10*np.log10(mfr_psd[i]))
    ax.set_yscale('log')
    ax.set(xlabel="Freq (1/min)",ylabel="Power Spectral Density (dB)",title=f"Power Spectral Density TT {rpm} % of 10RPM")
    f.savefig(f"fft_flowrate_TT_{rpm}_logscale.png")

def plot_all_fft():
    f,ax = plt.subplots(2,5)
    for ii,(rpm,v) in enumerate(data.items()):
        mfr_fft = fftpack.fft(v[:,2])
        # power spectral density
        mfr_psd = np.abs(v[:,2])**2
        # get frequencies correponding to PSD
        # setting frequency to 1 to 1 min
        fftfreq = fftpack.fftfreq(len(mfr_psd),1./60.)
        # only getting the positive parts as they correspond to the real signal
        i = fftfreq>0
        # plot
        ax[np.unravel_index(ii,ax.shape)].plot(fftfreq[i],10*np.log10(mfr_psd[i]))
        ax[np.unravel_index(ii,ax.shape)].set(xlabel="Freq (1/min)",ylabel="Power Spectral Density (dB)",title=f"Power Spectral Density for TT {rpm} RPM")

def plot_fft_rpm(rpm):
    # compute fourier transform
    mfr_fft = fftpack.fft(data[rpm][:,2])
    # power spectral density
    mfr_psd = np.abs(data[rpm][:,2])**2
    # get frequencies correponding to PSD
    # setting frequency to 1 to 1 min
    fftfreq = fftpack.fftfreq(len(mfr_psd),1./60.)
    # only getting the positive parts as they correspond to the real signal
    i = fftfreq>0
    f,ax = plt.subplots()
    ax.plot(fftfreq[i],10*np.log10(mfr_psd[i]))
    ax.set(xlabel="Freq (1/min)",ylabel="Power Spectral Density (dB)",title=f"Power Spectral Density for TT {rpm} RPM")

def plot_all_mfr():
    f,ax = plt.subplots(2,5,constrained_layout=True)
    for ii,(rpm,v) in enumerate(data.items()):
        ax[np.unravel_index(ii,ax.shape)].plot(data[rpm][:,1],data[rpm][:,2],'b-',data[rpm][:,1],data[rpm][:,3],'r-')
        ax[np.unravel_index(ii,ax.shape)].set(xlabel="Time (min)",ylabel="Flow Rate (g/min)",title=f"Flow Rate for TT Speed {rpm} RPM")
    f.legend(["Raw","Smoothed"])

def plot_fr_rpm(rpm):
    f,ax = plt.subplots()
    ax.plot(data[rpm][:,1],data[rpm][:,2],'b-',data[rpm][:,1],data[rpm][:,3],'r-')
    ax.set(xlabel="Time (min)",ylabel="Flow Rate (g/min)",title=f"Flow Rate for Turntable Speed {rpm} RPM")
    f.legend(["Column 3","Column 4"])
