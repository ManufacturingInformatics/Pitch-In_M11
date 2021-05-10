from scipy import fftpack
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import os
import pywt
import scaleogram as scg


def processMFRWavelets(fname,ng,gnames=None,opath='.',save_split_data=False,wavelet='morl'):
    if gnames is None:
        gnames = [f"Group_{i}" for i in range(ng)]
    # form output foldername
    opath = os.path.join(opath,os.path.splitext(os.path.basename(fname))[0])
    print(f"output path set to {opath}")
    # make the directory for the output
    os.makedirs(opath,exist_ok=True)
    # read in the workbook
    wb = load_workbook(filename=fname,data_only=True)
    print(f"workbook has sheets called {wb.sheetnames}")
    # get data worksheet
    print("accessing data worksheet")
    data_ws = wb["Data"]
    # data is broken up into collections of 5 columns with one columns separating each group
    # we're interested in the middle three
    # put data into an array
    #all_data= np.array([d for d in data_ws.iter_rows(min_row=8,max_col=60,values_only=True)])
    # dictionary to hold data
    data = {}
    # starting column index
    mc = 2
    # plotting axes
    f,ax = plt.subplots()
    # iterate over turnable speed groupings
    for ttg,gn in zip(range(1,ng+1),gnames):
        print(f"attempting to access {gn} data")
        print(f"col {mc} {mc+3}")
        # get data for the specific grouping
        dgroup = np.array([d for d in data_ws.iter_rows(min_row=8,min_col=mc,max_col=mc+3,values_only=True)])
        if save_split_data:
            # save subdataset
            np.savetxt(os.path.join(opath,f"{gn}_data.csv"),dgroup.astype("float64"),delimiter=',')
        # update column index
        mc += 6
        # add to dictionary
        data[gn]=dgroup.astype("float64")
        ## plot data
        # find where nans don't occur
        nn = ~ np.isnan(data[gn][:,2])
        ax.clear()
        scg.cws(data[gn][:,1][nn],data[gn][:,2][nn]-data[gn][:,2][nn].mean(),scales=np.arange(1,150),wavelet=wavelet,yscale='log',ax=ax,cmap='jet',cbar=None,ylabel="Period [mins]",xlabel="Time [mins]",title=f"Scaleogram for {gn} with log scale")
        f.savefig(os.path.join(opath,f"scaleogram_flowrate_TT_{gn}_logscale.png"))

        ax.clear()
        scg.cws(data[gn][:,1][nn],data[gn][:,2][nn]-data[gn][:,2][nn].mean(),scales=np.arange(1,150),wavelet=wavelet,ax=ax,cmap='jet',cbar=None,ylabel="Period [mins]",xlabel="Time [mins]",title=f"Scaleogram for {gn}")
        f.savefig(os.path.join(opath,f"scaleogram_flowrate_TT_{gn}.png"))


def processMFRFFT(fname,ng,gnames=None,opath='.',save_split_data=False):
    if gnames is None:
        gnames = [f"Group_{i}" for i in range(ng)]
    # form output foldername
    opath = os.path.join(opath,os.path.splitext(os.path.basename(fname))[0])
    print(f"output path set to {opath}")
    # make the directory for the output
    os.makedirs(opath,exist_ok=True)
    # read in the workbook
    wb = load_workbook(filename=fname,data_only=True)
    print(f"workbook has sheets called {wb.sheetnames}")
    # get data worksheet
    print("accessing data worksheet")
    data_ws = wb["Data"]
    # data is broken up into collections of 5 columns with one columns separating each group
    # we're interested in the middle three
    # put data into an array
    #all_data= np.array([d for d in data_ws.iter_rows(min_row=8,max_col=60,values_only=True)])
    # dictionary to hold data
    data = {}
    # starting column index
    mc = 2
    # plotting axes
    f,ax = plt.subplots()
    # iterate over turnable speed groupings
    for ttg,gn in zip(range(1,ng+1),gnames):
        print(f"attempting to access {gn} gname data")
        print(f"col {mc} {mc+3}")
        # get data for the specific grouping
        dgroup = np.array([d for d in data_ws.iter_rows(min_row=8,min_col=mc,max_col=mc+3,values_only=True)])
        if save_split_data:
            # save subdataset
            np.savetxt(os.path.join(opath,f"{gn}_data.csv"),dgroup.astype("float64"),delimiter=',')
        # update column index
        mc += 6
        # add to dictionary
        data[gn]=dgroup.astype("float64")
        # plot the data
        ax.clear()
        ax.plot(data[gn][:,1],data[gn][:,2],'b-',data[gn][:,1],data[gn][:,3],'r-')
        ax.set(xlabel="Time (min)",ylabel="Flow Rate (g/min)",title=f"Flow Rate for {gn}")
        leg = f.legend(["Raw","Smoothed"])
        f.savefig(os.path.join(opath,f"flowrate_TT_{gn}.png"))
        ## fft analysis
        print("performing fft analysis")
        # remove legend
        leg.remove()
        # compute fourier transform
        mfr_fft = fftpack.fft(data[gn][:,2])
        # power spectral density
        mfr_psd = np.abs(data[gn][:,2])**2
        # get frequencies correponding to PSD
        # setting frequency to 1 to 1 min
        fftfreq = fftpack.fftfreq(len(mfr_psd),1./60.)
        # only getting the positive parts as they correspond to the real signal
        i = fftfreq>0
        ## plot
        ax.clear()
        ax.plot(fftfreq[i],10*np.log10(mfr_psd[i]))
        ax.set(xlabel="Freq (1/min)",ylabel="Power Spectral Density (dB)",title=f"Power Spectral Density of {gn}")
        f.savefig(os.path.join(opath,f"fft_flowrate_TT_{gn}.png"))

        ax.clear()
        ax.plot(fftfreq[i],10*np.log10(mfr_psd[i]))
        ax.set_yscale('log')
        ax.set(xlabel="Freq (1/min)",ylabel="Power Spectral Density (dB)",title=f"Power Spectral Density {gn}")
        f.savefig(os.path.join(opath,f"fft_flowrate_TT_{gn}_logscale.png"))

def plot_all_fft():
    f,ax = plt.subplots(2,5)
    for ii,(rpm,v) in enumerate(data.items()):
        mfr_fft = fftpack.fft(v[:,2])
        # power spectral density
        mfr_psd = np.abs(v[:,2])**2
        # get frequencies correponding to PSD
        # setting frequency to 1 to 1 min
        fftfreq = fftpack.fftfreq(len(mfr_psd),1./60.)
        # only getting the positive parts as they correspond to the real signal
        i = fftfreq>0
        # plot
        ax[np.unravel_index(ii,ax.shape)].plot(fftfreq[i],10*np.log10(mfr_psd[i]))
        ax[np.unravel_index(ii,ax.shape)].set(xlabel="Freq (1/min)",ylabel="Power Spectral Density (dB)",title=f"Power Spectral Density for TT {rpm} RPM")

def plot_fft_rpm(rpm,data):
    # compute fourier transform
    mfr_fft = fftpack.fft(data[rpm][:,2])
    # power spectral density
    mfr_psd = np.abs(data[rpm][:,2])**2
    # get frequencies correponding to PSD
    # setting frequency to 1 to 1 min
    fftfreq = fftpack.fftfreq(len(mfr_psd),1./60.)
    # only getting the positive parts as they correspond to the real signal
    i = fftfreq>0
    f,ax = plt.subplots()
    ax.plot(fftfreq[i],10*np.log10(mfr_psd[i]))
    ax.set(xlabel="Freq (1/min)",ylabel="Power Spectral Density (dB)",title=f"Power Spectral Density for TT {rpm} RPM")

def plot_all_mfr():
    f,ax = plt.subplots(2,5,constrained_layout=True)
    for ii,(rpm,v) in enumerate(data.items()):
        ax[np.unravel_index(ii,ax.shape)].plot(data[rpm][:,1],data[rpm][:,2],'b-',data[rpm][:,1],data[rpm][:,3],'r-')
        ax[np.unravel_index(ii,ax.shape)].set(xlabel="Time (min)",ylabel="Flow Rate (g/min)",title=f"Flow Rate for TT Speed {rpm} RPM")
    f.legend(["Raw","Smoothed"])

def plot_fr_rpm(rpm,data):
    f,ax = plt.subplots()
    ax.plot(data[rpm][:,1],data[rpm][:,2],'b-',data[rpm][:,1],data[rpm][:,3],'r-')
    ax.set(xlabel="Time (min)",ylabel="Flow Rate (g/min)",title=f"Flow Rate for Turntable Speed {rpm} RPM")
    f.legend(["Column 3","Column 4"])
